import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def get_val(row, dict_key, orm_key):
    """
    Умный извлекатель данных. Понимает и словари, и объекты SQLAlchemy ORM.
    """
    if isinstance(row, dict):
        val = row.get(dict_key)
    else:
        val = getattr(row, orm_key, None)

    # Красивое форматирование дат
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y %H:%M")

    # Если это список (например, список препаратов), склеиваем в строку
    if isinstance(val, list):
        return ", ".join([str(i) for i in val])

    return val if val is not None else "—"


def create_excel_report(doc_data: list, apt_data: list) -> io.BytesIO:
    """
    Генерирует Excel файл с двумя листами: Врачи и Аптеки.
    """
    wb = Workbook()

    # ==========================================
    # 📄 ЛИСТ 1: ОТЧЕТЫ ПО ВРАЧАМ
    # ==========================================
    ws1 = wb.active
    ws1.title = "Врачи"

    headers1 = [
        "ID", "Дата", "Сотрудник",
        "Район", "Маршрут", "ЛПУ",
        "Врач", "Специальность", "Телефон",
        "Условия", "Препараты", "Комментарий"
    ]
    ws1.append(headers1)

    if doc_data:
        for row in doc_data:
            # Используем безопасное извлечение.
            # ORM ключи могут называться чуть иначе, добавил гибкость:
            user = get_val(row, 'user_name', 'user')
            date = get_val(row, 'created_at', 'date')
            comms = get_val(row, 'commentary', 'commentary') or get_val(row, 'comment', 'comment')

            ws1.append([
                get_val(row, 'id', 'id'),
                date,
                user,
                get_val(row, 'district', 'district'),
                get_val(row, 'road', 'road'),
                get_val(row, 'lpu', 'lpu'),
                get_val(row, 'doctor_name', 'doctor_name'),
                get_val(row, 'doctor_spec', 'doctor_spec'),
                get_val(row, 'doctor_number', 'doctor_number'),
                get_val(row, 'term', 'term'),
                get_val(row, 'preps', 'preps'),  # В БД это может быть строка или список
                comms
            ])

    # ==========================================
    # 📄 ЛИСТ 2: ОТЧЕТЫ ПО АПТЕКАМ
    # ==========================================
    ws2 = wb.create_sheet(title="Аптеки")

    headers2 = [
        "ID", "Дата", "Сотрудник",
        "Район", "Маршрут", "Точка (Аптека)",
        "Препарат", "Заявка (шт)", "Остаток (шт)",
        "Комментарий"
    ]
    ws2.append(headers2)

    if apt_data:
        for row in apt_data:
            user = get_val(row, 'user_name', 'user')
            date = get_val(row, 'created_at', 'date')
            lpu = get_val(row, 'lpu', 'lpu') or get_val(row, 'apothecary', 'apothecary')
            comms = get_val(row, 'commentary', 'commentary') or get_val(row, 'comment', 'comment')

            ws2.append([
                get_val(row, 'id', 'id'),
                date,
                user,
                get_val(row, 'district', 'district'),
                get_val(row, 'road', 'road'),
                lpu,
                get_val(row, 'prep_name', 'prep_name'),
                get_val(row, 'req_qty', 'req_qty'),
                get_val(row, 'rem_qty', 'rem_qty'),
                comms
            ])

    # ==========================================
    # 🎨 ОФОРМЛЕНИЕ (АВТО-ШИРИНА И ЦВЕТА)
    # ==========================================
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # ==========================================
    # 💾 СОХРАНЕНИЕ В ПАМЯТЬ
    # ==========================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output